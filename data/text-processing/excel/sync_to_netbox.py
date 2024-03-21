import re
import json
import requests
from openpyxl import load_workbook

base_url = "http://192.168.6.140:31868"
token = "0e8087028142a3d41d315dfd383e84e07135f9e3"


def extract_ip(text="") -> str:
    regex = r"(((25[0-5]|2[0-4][0-9]|[0-1]?[0-9][0-9]?)\.){3}(25[0-5]|2[0-4][0-9]|[0-1]?[0-9][0-9]?))"
    m = re.search(regex, text)
    if m:
        return m.group(0)
    else:
        # return "NO_IP_IN_NAME"
        return "0.0.0.0"


def request_api(api: str) -> dict:
    headers = {"Authorization": f"Token {token}"}

    response = requests.get(api, headers=headers)

    if response.status_code == 200:
        return response.json()
    else:
        print("Request failed:", response.status_code)
        return {}


def get_ip_list() -> list:
    wb = load_workbook(filename="application.xlsx")
    sheet = wb.active

    # count = 0
    # print(f"Count: {count}")

    selected_columns = [4]

    ips = []
    for row in sheet.iter_rows(values_only=True):
        if not all(cell is None or cell == "" for cell in row):
            data = row[3]
            # print(data)
            ips.append(data)

    wb.close()
    return list(set(ips))


def check_belonging(ip: str) -> tuple:
    device_tpl = f"{base_url}/api/dcim/devices/?q="
    vm_tpl = f"{base_url}/api/virtualization/virtual-machines/?q="

    device_names = []
    vm_names = []

    device_result = request_api(f"{device_tpl}{ip}")
    if device_result["count"] > 0:
        device_names = [item["name"] for item in device_result["results"]]
        # Only one match, check first
        if len(device_names) == 1 and (ip in device_names[0] or ip == device_names[0]):
            print(f"{ip} belongs to DEVICE: {device_names[0]}")
            return device_names[0], "device"

    vm_result = request_api(f"{vm_tpl}{ip}")
    if vm_result["count"] > 0:
        vm_names = [item["name"] for item in vm_result["results"]]
        # Only one match, check first
        if len(vm_names) == 1 and (ip in vm_names[0] or ip == vm_names[0]):
            print(f"{ip} belongs to VM: {vm_names[0]}")
            return vm_names[0], "vm"

    # print(f"device names: {device_names}")
    # print(f"vm names: {vm_names}")
    for name in device_names:
        ex_ip = extract_ip(name)
        if ip == ex_ip:
            print(f"{ip} belongs to device: {name}")
            return name, "device"

    for name in vm_names:
        ex_ip = extract_ip(name)
        if ip == ex_ip:
            print(f"{ip} belongs to vm: {name}")
            return name, "vm"

    print(f"{ip} not found")
    return None


if __name__ == "__main__":
    # get_ip_list()

    # parent_map = {}
    # for ip in get_ip_list():
    #     tmp = check_belonging(ip)
    #     if tmp:
    #         name, tp = tmp
    #         parent_map[ip] = {"type": tp, "name": name}
    #     else:
    #         parent_map[ip] = {}
    
    # with open("parent.json", "w") as f:
    #     json.dump(parent_map,fp=f)
    
    parent_map = {}
    with open('parent.json', 'r') as f:
        parent_map = json.load(f)


    success = 0
    failed = 0
    wb = load_workbook(filename="application.xlsx")
    sheet = wb.active
    all_rows = sheet.iter_rows(values_only=True)
    for row in all_rows:
        if not all(cell is None or cell == '' for cell in row):
            name, subsystem, access, ip, ports, path, runtime, db, comments = row[:9]
            # print(name, subsystem, access, ip, ports, path, runtime, db, comments)

            data = {
                    # "virtual_machine": {
                    #     "name": "10.0.102.10"
                    # },
                    # "device": {
                    #     "name": "10.1.66.10 安恒信息"
                    # },
                    "name": name if name else "未知服务",
                    "custom_fields": {
                        "Access": access,
                        "AppPath": path,
                        "Configuration": None,
                        "Runtime": runtime,
                        "SubSystem": subsystem,
                    }
                }
            parent = parent_map[ip]
            if parent['type'] == "vm":
                data['virtual_machine'] = {'name': parent['name'] }
            else:
                data['device'] = {'name': parent['name'] }
            

            data['protocol']="tcp"
            if ports:
                # print(f'------: {ports}, {name}, {subsystem}')
                if isinstance(ports, str):
                    l = re.split(r'[\n|]+',ports)
                    # print(ports.split('|'))
                    # print([lambda x: int(x.strip()) for x in ports.split('|')])
                    new_ports = []
                    for item in l:
                        if item:
                            port_no = int(item.strip())
                            if port_no > 65535:
                                continue
                            new_ports.append(int(item.strip()))

                    data['ports'] = new_ports
                else:
                    data['ports'] = [ports]
            else:
                data['ports'] = [1]


            
    
            url = f"{base_url}/api/ipam/services/"
            headers = {"Authorization": f"Token {token}"}

            response = requests.post(url, headers=headers, json=data)
            if response.status_code > 300:
                print(response.text)
                failed += 1
            else:
                print(response.status_code)
                success += 1
    
    print(f"success: {success}, failed: {failed}")